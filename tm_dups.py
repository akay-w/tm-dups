# -*- coding: utf-8 -*-
"""
Created on Mon Jul  1 17:37:25 2019

@author: a-whalen
"""

import openpyxl

file = r"" #insert path here
wb = openpyxl.load_workbook(file)
ws = wb["Sheet1"]

pairs = {}
max_row = ws.max_row
for i in range(2,max_row):
    jap = ws.cell(row = i, column = 4).value
    eng = ws.cell(row = i, column = 5).value
    attr = ws.cell(row = i, column = 6).value
    if jap and eng:
        if jap not in pairs.keys():
            pairs[jap] = [(eng, attr)]
        else:
            pairs[jap].append((eng, attr))
wb.close()

for k, v in pairs.items():
    print(k, v)
    
report = openpyxl.Workbook()
sheet = report.active
row = 2
col = 1
attr1 = "AK"
attr2 = "YG"
for jap, data in pairs.items():
    if len(data) > 1:
        for tup in data:
            eng = tup[0]
            attr = tup[1]
            sheet.cell(row = row, column = col).value = jap
            sheet.cell(row = row, column = col+1).value = eng
            if attr == None:
                sheet.cell(row = row, column = col+2).value = attr1
            elif attr2 in attr:
                sheet.cell(row = row, column = col+2).value = attr2
            else:
                sheet.cell(row = row, column = col+2).value = attr1
            row += 1
            
report.save("Report.xlsx")
report.close()

        
        
        